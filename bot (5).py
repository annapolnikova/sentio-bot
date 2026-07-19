"""
Sentio Test Group Bot
----------------------
Регистрирует участниц тестовой группы, ежедневно присылает анкету
(вопросы взяты из Google Form "Відгук по використанню продуктів Sentio"),
напоминает тем, кто не ответил, и позволяет админу (Anna) выгрузить
все ответы в Excel.

Запуск: см. README.md
"""

import logging
import os
import re
import sqlite3
from datetime import datetime, time as dtime
from zoneinfo import ZoneInfo

from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# ---------------------------------------------------------------------------
# Настройки
# ---------------------------------------------------------------------------

BOT_TOKEN = os.environ["BOT_TOKEN"]  # обязателен
ADMIN_IDS = [int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x.strip()]
TIMEZONE = ZoneInfo(os.environ.get("TZ_NAME", "Europe/Kyiv"))
REMINDER_HOUR = int(os.environ.get("REMINDER_HOUR", "20"))
REMINDER_MINUTE = int(os.environ.get("REMINDER_MINUTE", "0"))
DB_PATH = os.environ.get("DB_PATH", "sentio.db")
TEST_PERIOD_DAYS = int(os.environ.get("TEST_PERIOD_DAYS", "14"))

THANK_YOU_MESSAGE = (
    "Дякуємо, що була з нами ці 14 днів тестування Sentio! 🤍\n\n"
    "Твої відповіді дуже допомогли нам зробити продукт кращим. "
    "Анкети більше надсилатись не будуть.\n\n"
    "Якщо захочеш поділитися ще якимись враженнями — просто напиши нам прямо тут, "
    "ми завжди раді почути.\n\n"
    "З любов'ю, команда Sentio 🌿"
)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# База данных
# ---------------------------------------------------------------------------

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = db()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            telegram_id INTEGER PRIMARY KEY,
            username TEXT,
            name TEXT,
            email TEXT,
            phone TEXT,
            age_group TEXT,
            rollon TEXT,
            registered_at TEXT,
            completed_notified INTEGER DEFAULT 0,
            last_reminder_date TEXT
        );

        CREATE TABLE IF NOT EXISTS responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER,
            response_date TEXT,
            rollon_product TEXT,
            used_today TEXT,
            state_change TEXT,
            feelings TEXT,
            effect_time TEXT,
            aroma_rating TEXT,
            usage_freq TEXT,
            experience_text TEXT,
            will_continue TEXT,
            consent TEXT,
            instagram TEXT,
            created_at TEXT,
            FOREIGN KEY (telegram_id) REFERENCES users (telegram_id)
        );

        CREATE TABLE IF NOT EXISTS daily_feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER,
            response_date TEXT,
            consent TEXT,
            created_at TEXT,
            FOREIGN KEY (telegram_id) REFERENCES users (telegram_id)
        );
        """
    )
    conn.commit()
    conn.close()
    # На випадок, якщо база вже існувала без колонки rollon_product (оновлення бота)
    conn = db()
    cols = [r[1] for r in conn.execute("PRAGMA table_info(responses)").fetchall()]
    if "rollon_product" not in cols:
        conn.execute("ALTER TABLE responses ADD COLUMN rollon_product TEXT")
        conn.commit()
    if "used_today" not in cols:
        conn.execute("ALTER TABLE responses ADD COLUMN used_today TEXT")
        conn.commit()
    conn.close()
    conn = db()
    ucols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if "completed_notified" not in ucols:
        conn.execute("ALTER TABLE users ADD COLUMN completed_notified INTEGER DEFAULT 0")
        conn.commit()
    if "last_reminder_date" not in ucols:
        conn.execute("ALTER TABLE users ADD COLUMN last_reminder_date TEXT")
        conn.commit()
    conn.close()


def days_since_registration(user: dict) -> int:
    reg_date = datetime.fromisoformat(user["registered_at"]).astimezone(TIMEZONE).date()
    today = datetime.now(TIMEZONE).date()
    return (today - reg_date).days


def mark_completed_notified(telegram_id: int):
    conn = db()
    conn.execute(
        "UPDATE users SET completed_notified = 1 WHERE telegram_id = ?", (telegram_id,)
    )
    conn.commit()
    conn.close()


def user_exists(telegram_id: int) -> bool:
    conn = db()
    row = conn.execute(
        "SELECT 1 FROM users WHERE telegram_id = ?", (telegram_id,)
    ).fetchone()
    conn.close()
    return row is not None


def get_user(telegram_id: int):
    conn = db()
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT * FROM users WHERE telegram_id = ?", (telegram_id,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def save_user(data: dict):
    existing = get_user(data["telegram_id"])
    registered_at = existing["registered_at"] if existing else datetime.now(TIMEZONE).isoformat()
    completed_notified = existing["completed_notified"] if existing else 0
    last_reminder_date = existing["last_reminder_date"] if existing else None
    conn = db()
    conn.execute(
        """INSERT OR REPLACE INTO users
           (telegram_id, username, name, email, phone, age_group, rollon, registered_at, completed_notified, last_reminder_date)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            data["telegram_id"],
            data.get("username", ""),
            data["name"],
            data["email"],
            data["phone"],
            data["age_group"],
            data["rollon"],
            registered_at,
            completed_notified,
            last_reminder_date,
        ),
    )
    conn.commit()
    conn.close()


def mark_reminder_sent(telegram_id: int):
    today = datetime.now(TIMEZONE).date().isoformat()
    conn = db()
    conn.execute(
        "UPDATE users SET last_reminder_date = ? WHERE telegram_id = ?", (today, telegram_id)
    )
    conn.commit()
    conn.close()


def all_user_ids():
    conn = db()
    rows = conn.execute("SELECT telegram_id FROM users").fetchall()
    conn.close()
    return [r[0] for r in rows]


def already_responded_today(telegram_id: int, product: str) -> bool:
    today = datetime.now(TIMEZONE).date().isoformat()
    conn = db()
    row = conn.execute(
        "SELECT 1 FROM responses WHERE telegram_id = ? AND response_date = ? AND rollon_product = ?",
        (telegram_id, today, product),
    ).fetchone()
    conn.close()
    return row is not None


def has_pending_products_today(telegram_id: int) -> bool:
    user = get_user(telegram_id)
    if not user:
        return False
    products = ROLLON_TO_PRODUCTS.get(user.get("rollon"), ["RESET"])
    return any(not already_responded_today(telegram_id, p) for p in products)


def save_response(telegram_id: int, product: str, answers: dict):
    conn = db()
    today = datetime.now(TIMEZONE).date().isoformat()
    conn.execute(
        """INSERT INTO responses
           (telegram_id, response_date, rollon_product, used_today, state_change, feelings, effect_time,
            aroma_rating, usage_freq, experience_text, will_continue, consent,
            instagram, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            telegram_id,
            today,
            product,
            answers.get("used_today", "так"),
            answers.get("state_change", ""),
            answers.get("feelings", ""),
            answers.get("effect_time", ""),
            answers.get("aroma_rating", ""),
            answers.get("usage_freq", ""),
            answers.get("experience_text", ""),
            answers.get("will_continue", ""),
            answers.get("consent", ""),
            answers.get("instagram", ""),
            datetime.now(TIMEZONE).isoformat(),
        ),
    )
    conn.commit()
    conn.close()


def save_daily_consent(telegram_id: int, consent: str):
    conn = db()
    today = datetime.now(TIMEZONE).date().isoformat()
    conn.execute(
        "INSERT INTO daily_feedback (telegram_id, response_date, consent, created_at) VALUES (?, ?, ?, ?)",
        (telegram_id, today, consent, datetime.now(TIMEZONE).isoformat()),
    )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Регистрация — состояния
# ---------------------------------------------------------------------------

REG_NAME, REG_EMAIL, REG_PHONE, REG_AGE, REG_ROLLON = range(5)

AGE_OPTIONS = ["до 25", "25–34", "35–44", "45+"]
ROLLON_OPTIONS = [
    "Всі три",
    "Reset - no panic (анти-тривога)",
    "Rise - no noise / morning (фокус / ясність)",
    "Rest - no rush (сон / розслаблення)",
]

ROLLON_TO_PRODUCTS = {
    "Всі три": ["RESET", "RISE", "REST"],
    "Reset - no panic (анти-тривога)": ["RESET"],
    "Rise - no noise / morning (фокус / ясність)": ["RISE"],
    "Rest - no rush (сон / розслаблення)": ["REST"],
}
PRODUCT_LABELS = {
    "RESET": "Reset – no panic (анти-тривога)",
    "RISE": "Rise – no noise / morning (фокус / ясність)",
    "REST": "Rest – no rush (сон / розслаблення)",
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")


def kb(options, prefix):
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton(o, callback_data=f"{prefix}:{o}")] for o in options]
    )


def kb_with_current(options, prefix, current):
    buttons = []
    if current and current in options:
        buttons.append([InlineKeyboardButton(f"✅ Залишити: {current}", callback_data=f"{prefix}:{current}")])
    buttons += [[InlineKeyboardButton(o, callback_data=f"{prefix}:{o}")] for o in options if o != current]
    return InlineKeyboardMarkup(buttons)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if user_exists(tg_id):
        user = get_user(tg_id)
        if days_since_registration(user) >= TEST_PERIOD_DAYS:
            if not user.get("completed_notified"):
                mark_completed_notified(tg_id)
            await update.message.reply_text(THANK_YOU_MESSAGE)
            return ConversationHandler.END

        markup = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("📝 Заповнити анкету", callback_data="goto_survey")],
                [InlineKeyboardButton("✏️ Оновити дані реєстрації", callback_data="start_update")],
            ]
        )
        await update.message.reply_text(
            "Ти вже зареєстрована в тестовій групі Sentio 🤍\n"
            "Щодня о 20:00 я нагадаю заповнити анкету.",
            reply_markup=markup,
        )
        return ConversationHandler.END

    context.user_data["updating"] = False
    await update.message.reply_text(
        "Привіт! 🤍 Це тестова група Sentio.\n"
        "Давай зареєструємось — це займе хвилину.\n\n"
        "Як тебе звати?",
        reply_markup=ReplyKeyboardRemove(),
    )
    return REG_NAME


async def start_update(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id
    old = get_user(tg_id) or {}
    context.user_data["updating"] = True
    context.user_data["old"] = old
    await q.edit_message_text("Оновлюємо дані реєстрації.")
    await q.message.reply_text(
        f"Поточне ім'я: {old.get('name', '—')}\n"
        "Введи нове ім'я, або надішли крапку «.», щоб залишити без змін.",
        reply_markup=ReplyKeyboardRemove(),
    )
    return REG_NAME


async def reg_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    old = context.user_data.get("old", {})
    if context.user_data.get("updating") and text == ".":
        context.user_data["name"] = old.get("name", "")
    else:
        context.user_data["name"] = text

    if context.user_data.get("updating"):
        await update.message.reply_text(
            f"Поточний e-mail: {old.get('email', '—')}\n"
            "Введи новий e-mail, або надішли крапку «.», щоб залишити без змін."
        )
    else:
        await update.message.reply_text("Твій e-mail?")
    return REG_EMAIL


async def reg_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    old = context.user_data.get("old", {})
    if context.user_data.get("updating") and text == ".":
        context.user_data["email"] = old.get("email", "")
    else:
        if not EMAIL_RE.match(text):
            await update.message.reply_text(
                "Схоже, це не e-mail 🤔 Приклад правильного формату: name@example.com\n"
                "Спробуй ще раз:"
            )
            return REG_EMAIL
        context.user_data["email"] = text

    phone_kb_buttons = [[KeyboardButton("📱 Поділитися номером", request_contact=True)]]
    phone_kb = ReplyKeyboardMarkup(phone_kb_buttons, resize_keyboard=True, one_time_keyboard=True)
    if context.user_data.get("updating"):
        await update.message.reply_text(
            f"Поточний телефон: {old.get('phone', '—')}\n"
            "Надішли новий (кнопкою або текстом), або крапку «.», щоб залишити без змін.",
            reply_markup=phone_kb,
        )
    else:
        await update.message.reply_text(
            "Твій телефон? Можеш натиснути кнопку нижче або ввести вручну.",
            reply_markup=phone_kb,
        )
    return REG_PHONE


async def reg_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    old = context.user_data.get("old", {})
    if update.message.contact:
        context.user_data["phone"] = update.message.contact.phone_number
    else:
        text = update.message.text.strip()
        if context.user_data.get("updating") and text == ".":
            context.user_data["phone"] = old.get("phone", "")
        else:
            context.user_data["phone"] = text

    await update.message.reply_text("Твій вік?", reply_markup=ReplyKeyboardRemove())
    age_kb = kb_with_current(AGE_OPTIONS, "age", old.get("age_group")) if context.user_data.get("updating") else kb(AGE_OPTIONS, "age")
    await update.message.reply_text("Обери варіант:", reply_markup=age_kb)
    return REG_AGE


async def reg_age(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["age_group"] = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Вік: {context.user_data['age_group']}")
    old = context.user_data.get("old", {})
    rollon_kb = kb_with_current(ROLLON_OPTIONS, "rollon", old.get("rollon")) if context.user_data.get("updating") else kb(ROLLON_OPTIONS, "rollon")
    await q.message.reply_text("Який рол-он ти тестуєш?", reply_markup=rollon_kb)
    return REG_ROLLON


async def reg_rollon(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["rollon"] = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Рол-он: {context.user_data['rollon']}")

    data = {
        "telegram_id": update.effective_user.id,
        "username": update.effective_user.username or "",
        "name": context.user_data["name"],
        "email": context.user_data["email"],
        "phone": context.user_data["phone"],
        "age_group": context.user_data["age_group"],
        "rollon": context.user_data["rollon"],
    }
    save_user(data)
    was_updating = context.user_data.get("updating")

    if was_updating:
        await q.message.reply_text("Дані оновлено! 🤍 Можеш заповнити анкету командою /survey")
    else:
        await q.message.reply_text(
            "Дякуємо, реєстрацію завершено! 🤍\n"
            f"Щодня о {REMINDER_HOUR:02d}:{REMINDER_MINUTE:02d} я надсилатиму коротку анкету "
            "про твої відчуття від рол-она. Це займе 1-2 хвилини.\n\n"
            "Можеш заповнити першу анкету прямо зараз командою /survey"
        )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Скасовано. Напиши /start щоб почати знову.")
    return ConversationHandler.END


registration_conv = ConversationHandler(
    entry_points=[
        CommandHandler("start", start),
        CallbackQueryHandler(start_update, pattern=r"^start_update$"),
    ],
    states={
        REG_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_name)],
        REG_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_email)],
        REG_PHONE: [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), reg_phone)],
        REG_AGE: [CallbackQueryHandler(reg_age, pattern=r"^age:")],
        REG_ROLLON: [CallbackQueryHandler(reg_rollon, pattern=r"^rollon:")],
    },
    fallbacks=[
        CommandHandler("cancel", cancel),
        CommandHandler("start", start),
        CallbackQueryHandler(start_update, pattern=r"^start_update$"),
    ],
    conversation_timeout=7200,
)

# ---------------------------------------------------------------------------
# Ежедневная анкета — состояния
# ---------------------------------------------------------------------------

(
    SUR_USED,
    SUR_STATE,
    SUR_FEEL,
    SUR_TIME,
    SUR_AROMA,
    SUR_TEXT,
    SUR_CONTINUE,
    SUR_DAILY_CONSENT,
) = range(10, 18)

STATE_OPTIONS = ["без змін", "1", "2", "3", "4", "5 (сильний ефект)"]
FEEL_OPTIONS = [
    "заспокоєння",
    "менше тривоги",
    "легше заснути",
    "кращий настрій",
    "більше енергії",
    "легше зосередитись",
    "не відчула ефекту",
]
TIME_OPTIONS = ["1–3 хв", "5–10 хв", "10–20 хв", "не відчула"]
AROMA_OPTIONS = ["дуже подобається", "приємний", "нейтральний", "не сподобався"]
CONTINUE_OPTIONS = ["так", "можливо", "ні"]
CONSENT_OPTIONS = ["так", "так, без імені", "ні"]  # питається один раз на день, після всіх продуктів

USED_TODAY_KB = InlineKeyboardMarkup(
    [
        [InlineKeyboardButton("✅ Так, використовувала", callback_data="used:так")],
        [InlineKeyboardButton("🚫 Ні, не використовувала сьогодні", callback_data="used:ні")],
    ]
)


async def survey_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    user = get_user(tg_id)
    target = update.message or update.callback_query.message
    if update.callback_query:
        await update.callback_query.answer()
    if not user:
        await target.reply_text("Спочатку потрібно зареєструватись — напиши /start")
        return ConversationHandler.END

    if days_since_registration(user) >= TEST_PERIOD_DAYS:
        if not user.get("completed_notified"):
            mark_completed_notified(tg_id)
        await target.reply_text(THANK_YOU_MESSAGE)
        return ConversationHandler.END

    products = ROLLON_TO_PRODUCTS.get(user.get("rollon"), ["RESET"])
    remaining = [p for p in products if not already_responded_today(tg_id, p)]
    if not remaining:
        await target.reply_text("Ти вже заповнила анкету на сьогодні по всіх продуктах 🤍 Дякуємо!")
        return ConversationHandler.END

    current = remaining[0]
    context.user_data["survey"] = {"feelings": [], "product": current}
    context.user_data["survey_remaining"] = remaining[1:]

    intro = f"📝 Анкета по продукту: {PRODUCT_LABELS[current]}\n\n" if len(products) > 1 else ""
    await target.reply_text(
        intro + "Чи використовувала сьогодні цей рол-он?", reply_markup=USED_TODAY_KB
    )
    return SUR_USED


async def continue_or_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    remaining = context.user_data.get("survey_remaining", [])
    if remaining:
        next_product = remaining[0]
        context.user_data["survey"] = {"feelings": [], "product": next_product}
        context.user_data["survey_remaining"] = remaining[1:]
        await update.effective_message.reply_text(
            f"Дякую! 🤍 Тепер анкета по продукту: {PRODUCT_LABELS[next_product]}\n\n"
            "Чи використовувала сьогодні цей рол-он?",
            reply_markup=USED_TODAY_KB,
        )
        return SUR_USED

    await update.effective_message.reply_text(
        "Останнє питання на сьогодні:\nЧи можна використати твій відгук?",
        reply_markup=kb(CONSENT_OPTIONS, "dailyconsent"),
    )
    return SUR_DAILY_CONSENT


async def sur_daily_consent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    consent = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Згода: {consent}")
    save_daily_consent(update.effective_user.id, consent)
    await q.message.reply_text("Дякуємо за відповіді! 🤍 Побачимось завтра.")
    context.user_data.clear()
    return ConversationHandler.END


async def sur_used_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    choice = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Використовувала сьогодні: {choice}")

    if choice == "ні":
        product = context.user_data["survey"].pop("product")
        save_response(update.effective_user.id, product, {"used_today": "ні"})
        return await continue_or_finish(update, context)

    await q.message.reply_text("Як змінився твій стан?", reply_markup=kb(STATE_OPTIONS, "state"))
    return SUR_STATE


async def sur_state(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["survey"]["state_change"] = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Стан: {context.user_data['survey']['state_change']}")

    kb_feel = InlineKeyboardMarkup(
        [[InlineKeyboardButton(o, callback_data=f"feel:{o}")] for o in FEEL_OPTIONS]
        + [[InlineKeyboardButton("✅ Готово", callback_data="feel:DONE")]]
    )
    await q.message.reply_text(
        "Що ти відчула? (обери одне чи декілька, потім натисни Готово)",
        reply_markup=kb_feel,
    )
    return SUR_FEEL


async def sur_feel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    choice = q.data.split(":", 1)[1]
    selected = context.user_data["survey"]["feelings"]

    if choice == "DONE":
        text = ", ".join(selected) if selected else "(нічого не обрано)"
        await q.edit_message_text(f"Відчуття: {text}")
        await q.message.reply_text(
            "Через скільки часу з'явився ефект?", reply_markup=kb(TIME_OPTIONS, "time")
        )
        return SUR_TIME

    if choice in selected:
        selected.remove(choice)
    else:
        selected.append(choice)

    marked = [("☑ " if o in selected else "") + o for o in FEEL_OPTIONS]
    kb_feel = InlineKeyboardMarkup(
        [[InlineKeyboardButton(m, callback_data=f"feel:{o}")] for m, o in zip(marked, FEEL_OPTIONS)]
        + [[InlineKeyboardButton("✅ Готово", callback_data="feel:DONE")]]
    )
    await q.edit_message_reply_markup(reply_markup=kb_feel)
    return SUR_FEEL


async def sur_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["survey"]["effect_time"] = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Час ефекту: {context.user_data['survey']['effect_time']}")
    await q.message.reply_text("Як тобі аромат?", reply_markup=kb(AROMA_OPTIONS, "aroma"))
    return SUR_AROMA


async def sur_aroma(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["survey"]["aroma_rating"] = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Аромат: {context.user_data['survey']['aroma_rating']}")
    await q.message.reply_text("Опиши свій досвід своїми словами 🤍 (або напиши /skip)")
    return SUR_TEXT


async def sur_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["survey"]["experience_text"] = update.message.text.strip()
    await update.message.reply_text(
        "Чи будеш ти користуватись далі?", reply_markup=kb(CONTINUE_OPTIONS, "cont")
    )
    return SUR_CONTINUE


async def sur_text_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["survey"]["experience_text"] = ""
    await update.message.reply_text(
        "Чи будеш ти користуватись далі?", reply_markup=kb(CONTINUE_OPTIONS, "cont")
    )
    return SUR_CONTINUE


async def sur_continue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["survey"]["will_continue"] = q.data.split(":", 1)[1]
    await q.edit_message_text(f"Продовжить: {context.user_data['survey']['will_continue']}")
    return await finalize_survey(update, context)


async def finalize_survey(update: Update, context: ContextTypes.DEFAULT_TYPE):
    survey = context.user_data["survey"]
    survey["instagram"] = ""
    survey["feelings"] = ", ".join(survey.get("feelings", []))
    survey["used_today"] = "так"
    product = survey.pop("product")
    save_response(update.effective_user.id, product, survey)
    return await continue_or_finish(update, context)


survey_conv = ConversationHandler(
    entry_points=[
        CommandHandler("survey", survey_entry),
        CallbackQueryHandler(survey_entry, pattern=r"^(daily_survey|goto_survey)$"),
    ],
    states={
        SUR_USED: [CallbackQueryHandler(sur_used_choice, pattern=r"^used:")],
        SUR_STATE: [CallbackQueryHandler(sur_state, pattern=r"^state:")],
        SUR_FEEL: [CallbackQueryHandler(sur_feel, pattern=r"^feel:")],
        SUR_TIME: [CallbackQueryHandler(sur_time, pattern=r"^time:")],
        SUR_AROMA: [CallbackQueryHandler(sur_aroma, pattern=r"^aroma:")],
        SUR_TEXT: [
            CommandHandler("skip", sur_text_skip),
            MessageHandler(filters.TEXT & ~filters.COMMAND, sur_text),
        ],
        SUR_CONTINUE: [CallbackQueryHandler(sur_continue, pattern=r"^cont:")],
        SUR_DAILY_CONSENT: [CallbackQueryHandler(sur_daily_consent, pattern=r"^dailyconsent:")],
    },
    fallbacks=[
        CommandHandler("cancel", cancel),
        # Якщо людина застрягла на попередньому питанні і тисне кнопку нового
        # нагадування (або пише /survey ще раз) — просто перезапускаємо анкету.
        CommandHandler("survey", survey_entry),
        CallbackQueryHandler(survey_entry, pattern=r"^(daily_survey|goto_survey)$"),
    ],
    conversation_timeout=7200,  # 2 години — якщо анкету покинули на півдорозі, стан скидається сам
)

# ---------------------------------------------------------------------------
# Напоминание (ежедневная джоба)
# ---------------------------------------------------------------------------

async def send_daily_reminders(context: ContextTypes.DEFAULT_TYPE):
    today = datetime.now(TIMEZONE).date().isoformat()
    for tg_id in all_user_ids():
        user = get_user(tg_id)
        if not user:
            continue

        if days_since_registration(user) >= TEST_PERIOD_DAYS:
            if not user.get("completed_notified"):
                try:
                    await context.bot.send_message(chat_id=tg_id, text=THANK_YOU_MESSAGE)
                except Exception as e:
                    logger.warning("Не вдалось надіслати фінальне повідомлення %s: %s", tg_id, e)
                mark_completed_notified(tg_id)
            continue

        if user.get("last_reminder_date") == today:
            continue  # нагадування вже надсилалось сьогодні (захист від дублів при повторному запуску)

        if not has_pending_products_today(tg_id):
            continue
        try:
            markup = InlineKeyboardMarkup(
                [[InlineKeyboardButton("📝 Заповнити анкету", callback_data="daily_survey")]]
            )
            await context.bot.send_message(
                chat_id=tg_id,
                text="Привіт! 🌙 Час поділитись відчуттями від рол-она за сьогодні.",
                reply_markup=markup,
            )
            mark_reminder_sent(tg_id)
        except Exception as e:
            logger.warning("Не вдалось надіслати нагадування %s: %s", tg_id, e)


# ---------------------------------------------------------------------------
# Админ: экспорт и статистика
# ---------------------------------------------------------------------------

async def export_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("Ця команда доступна лише адміну.")
        return

    from openpyxl import Workbook

    conn = db()
    users = conn.execute("SELECT * FROM users").fetchall()
    user_cols = [d[0] for d in conn.execute("SELECT * FROM users").description]
    responses = conn.execute(
        "SELECT * FROM responses ORDER BY telegram_id, response_date"
    ).fetchall()
    resp_cols = [d[0] for d in conn.execute("SELECT * FROM responses").description]
    consents = conn.execute(
        "SELECT * FROM daily_feedback ORDER BY telegram_id, response_date"
    ).fetchall()
    consent_cols = [d[0] for d in conn.execute("SELECT * FROM daily_feedback").description]
    conn.close()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Users"
    ws1.append(user_cols)
    for row in users:
        ws1.append(row)

    ws2 = wb.create_sheet("Responses")
    ws2.append(resp_cols)
    for row in responses:
        ws2.append(row)

    ws3 = wb.create_sheet("DailyConsent")
    ws3.append(consent_cols)
    for row in consents:
        ws3.append(row)

    fname = f"sentio_export_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    await update.message.reply_document(document=open(fname, "rb"), filename=fname)
    os.remove(fname)


async def stats_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("Ця команда доступна лише адміну.")
        return
    conn = db()
    n_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    n_resp = conn.execute("SELECT COUNT(*) FROM responses").fetchone()[0]
    today = datetime.now(TIMEZONE).date().isoformat()
    n_today = conn.execute(
        "SELECT COUNT(*) FROM responses WHERE response_date = ?", (today,)
    ).fetchone()[0]
    conn.close()
    await update.message.reply_text(
        f"👥 Зареєстровано: {n_users}\n"
        f"📝 Всього відповідей: {n_resp}\n"
        f"📅 Відповіли сьогодні: {n_today}"
    )


async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Твій Telegram ID: {update.effective_user.id}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(registration_conv)
    app.add_handler(survey_conv)
    app.add_handler(CommandHandler("export", export_cmd))
    app.add_handler(CommandHandler("stats", stats_cmd))
    app.add_handler(CommandHandler("whoami", whoami))

    app.job_queue.run_daily(
        send_daily_reminders,
        time=dtime(hour=REMINDER_HOUR, minute=REMINDER_MINUTE, tzinfo=TIMEZONE),
    )

    # "Наздоганяюча" перевірка одразу після старту/передеплою: якщо сервіс
    # перезапустився вже ПІСЛЯ часу нагадування, а сьогоднішня розсилка ще не
    # пішла — надсилаємо її зараз, а не чекаємо до завтра.
    now = datetime.now(TIMEZONE)
    reminder_time_today = now.replace(
        hour=REMINDER_HOUR, minute=REMINDER_MINUTE, second=0, microsecond=0
    )
    if now >= reminder_time_today:
        app.job_queue.run_once(send_daily_reminders, when=15)

    logger.info("Bot started")
    app.run_polling()


if __name__ == "__main__":
    main()
